"""Unit tests for the in-place Override-CPC workbook writer.

These exercise the pure exporter directly (no database): given a workbook and a
set of updates, only the targeted Override CPC cells may change and every other
aspect of the workbook must survive a round-trip untouched.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from app.optimizer.excel_export import BidUpdate, BidWorkbookExporter, optimized_filename

PROVIDER = "CheapTicketsDealM_FIOAD_US"

# (provider, origin, destination, excluded, override_cpc) — provider only on row 1,
# matching the real KAYAK export (blank rows fall back to the Mode-sheet provider).
_ROWS: list[tuple[str | None, str, str, str, float]] = [
    (PROVIDER, "JFK", "LAX", "false", 1.60),
    (None, "JFK", "SFO", "false", 1.40),
    (None, "BOS", "MIA", "false", 1.80),
]


def _rich_workbook_bytes() -> bytes:
    """A bid workbook exercising every 'must preserve' feature."""
    wb = Workbook()
    st = wb.active
    st.title = "Search Terms"
    st.append(["Provider Code", "Origin", "Destination", "Excluded", "Override CPC"])
    for provider, origin, destination, excluded, cpc in _ROWS:
        st.append([provider, origin, destination, excluded, cpc])

    # Header formatting: bold white font, blue fill, thin border.
    thin = Side(style="thin", color="000000")
    for cell in st[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Number format on the Override CPC cells.
    for row in range(2, 2 + len(_ROWS)):
        st.cell(row=row, column=5).number_format = "0.00"
    # Column width, row height, freeze panes, merged title.
    st.column_dimensions["B"].width = 22.5
    st.row_dimensions[1].height = 30
    st.freeze_panes = "A2"
    st["G1"] = "Notes"
    st.merge_cells("G1:H1")
    # Data validation on the Excluded column.
    dv = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
    st.add_data_validation(dv)
    dv.add("D2:D4")

    mode = wb.create_sheet("Mode")
    mode.append(["Mode", "Full"])
    mode.append(["Provider Code", PROVIDER])

    # A second data sheet with a formula referencing the Override CPC column.
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Total"
    summary["B1"] = "=SUM('Search Terms'!E2:E4)"

    hidden = wb.create_sheet("Hidden")
    hidden["A1"] = "secret"
    hidden.sheet_state = "hidden"

    wb.defined_names["CpcRange"] = DefinedName("CpcRange", attr_text="'Search Terms'!$E$2:$E$4")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _search_terms(content: bytes) -> Workbook:
    return load_workbook(BytesIO(content))["Search Terms"]


# --- Filename ---------------------------------------------------------------


@pytest.mark.parametrize(
    ("original", "expected"),
    [
        ("CheapTicketsDealM_FIOAD_US.xlsx", "CheapTicketsDealM_FIOAD_US_Optimized.xlsx"),
        ("CheapTicketsDealD_FIOAD_US.xlsx", "CheapTicketsDealD_FIOAD_US_Optimized.xlsx"),
        ("bids.XLSX", "bids_Optimized.xlsx"),
        ("no_extension", "no_extension_Optimized.xlsx"),
    ],
)
def test_optimized_filename(original: str, expected: str) -> None:
    assert optimized_filename(original) == expected


# --- Override CPC updates ---------------------------------------------------


def test_increase_writes_recommended_bid() -> None:
    result = BidWorkbookExporter().apply_updates(
        _rich_workbook_bytes(), [BidUpdate(PROVIDER, "JFK", "LAX", 1.76)]
    )
    assert result.rows_updated == 1
    assert result.unmatched_updates == []
    st = _search_terms(result.content)
    assert st["E2"].value == 1.76  # JFK-LAX updated
    assert st["E3"].value == 1.40  # JFK-SFO untouched
    assert st["E4"].value == 1.80  # BOS-MIA untouched


def test_matches_blank_provider_rows_and_is_case_insensitive() -> None:
    # BOS-MIA is a blank-provider row; lower-case input must still match.
    result = BidWorkbookExporter().apply_updates(
        _rich_workbook_bytes(), [BidUpdate(PROVIDER, "bos", "mia", 2.10)]
    )
    assert result.rows_updated == 1
    assert _search_terms(result.content)["E4"].value == 2.10


def test_only_override_cpc_cells_change() -> None:
    source = _rich_workbook_bytes()
    result = BidWorkbookExporter().apply_updates(source, [BidUpdate(PROVIDER, "JFK", "LAX", 1.76)])
    before = _search_terms(source)
    after = _search_terms(result.content)
    for row in range(1, before.max_row + 1):
        for col in range(1, before.max_column + 1):
            b = before.cell(row=row, column=col).value
            a = after.cell(row=row, column=col).value
            if row == 2 and col == 5:  # the single updated Override CPC cell
                assert b == 1.60 and a == 1.76
            else:
                assert a == b, f"cell ({row},{col}) changed: {b!r} -> {a!r}"


def test_missing_route_is_recorded_not_fatal() -> None:
    source = _rich_workbook_bytes()
    result = BidWorkbookExporter().apply_updates(source, [BidUpdate(PROVIDER, "ZZZ", "YYY", 2.00)])
    assert result.rows_updated == 0
    assert len(result.unmatched_updates) == 1
    assert (result.unmatched_updates[0].origin, result.unmatched_updates[0].destination) == (
        "ZZZ",
        "YYY",
    )
    # Nothing in the workbook changed.
    assert _search_terms(result.content)["E2"].value == 1.60


def test_no_updates_leaves_workbook_intact() -> None:
    source = _rich_workbook_bytes()
    result = BidWorkbookExporter().apply_updates(source, [])
    assert result.rows_updated == 0
    before, after = _search_terms(source), _search_terms(result.content)
    for row in range(1, before.max_row + 1):
        for col in range(1, before.max_column + 1):
            assert after.cell(row=row, column=col).value == before.cell(row=row, column=col).value


# --- Preservation -----------------------------------------------------------


def test_worksheets_names_order_and_hidden_preserved() -> None:
    result = BidWorkbookExporter().apply_updates(
        _rich_workbook_bytes(), [BidUpdate(PROVIDER, "JFK", "LAX", 1.76)]
    )
    wb = load_workbook(BytesIO(result.content))
    assert wb.sheetnames == ["Search Terms", "Mode", "Summary", "Hidden"]
    assert wb["Hidden"].sheet_state == "hidden"
    assert wb["Mode"]["B2"].value == PROVIDER


def test_formatting_preserved() -> None:
    result = BidWorkbookExporter().apply_updates(
        _rich_workbook_bytes(), [BidUpdate(PROVIDER, "JFK", "LAX", 1.76)]
    )
    st = load_workbook(BytesIO(result.content))["Search Terms"]
    # Freeze panes, merged range, column width, row height.
    assert st.freeze_panes == "A2"
    assert "G1:H1" in {str(r) for r in st.merged_cells.ranges}
    assert st.column_dimensions["B"].width == 22.5
    assert st.row_dimensions[1].height == 30
    # Header font + fill.
    assert st["A1"].font.bold is True
    assert str(st["A1"].fill.fgColor.rgb).endswith("4472C4")
    # Number format survives on the cell that WAS updated.
    assert st["E2"].number_format == "0.00"
    # Data validation + named range survive.
    assert len(st.data_validations.dataValidation) == 1
    assert "CpcRange" in load_workbook(BytesIO(result.content)).defined_names


def test_formula_preserved() -> None:
    result = BidWorkbookExporter().apply_updates(
        _rich_workbook_bytes(), [BidUpdate(PROVIDER, "JFK", "LAX", 1.76)]
    )
    # Default load keeps formula strings (data_only=False).
    wb = load_workbook(BytesIO(result.content))
    assert wb["Summary"]["B1"].value == "=SUM('Search Terms'!E2:E4)"
