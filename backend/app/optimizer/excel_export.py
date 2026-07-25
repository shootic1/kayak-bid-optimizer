"""In-place Override-CPC writer for KAYAK bid workbooks.

Pure workbook logic: given the original uploaded workbook bytes and a set of
Override-CPC updates keyed by ``(provider, origin, destination)``, this produces
a new workbook that is **byte-for-byte identical except for the Override CPC
cells that changed**. It never edits any other cell, column, sheet, style,
formula, or piece of workbook metadata.

The engine that decides *which* routes to update lives elsewhere (Phase 3C);
this module only applies decisions. No database or HTTP concerns here.
"""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.importers.base import ParseError

SEARCH_TERMS_SHEET = "Search Terms"
MODE_SHEET = "Mode"

_EXPORT_SUFFIX = "_Optimized"
_XLSX_EXT = ".xlsx"


@dataclass(frozen=True)
class BidUpdate:
    """A single Override-CPC change to apply, addressed by provider + route."""

    provider_code: str
    origin: str
    destination: str
    new_cpc: float


@dataclass(frozen=True)
class WorkbookExportResult:
    """The rewritten workbook plus a record of what was (and wasn't) applied."""

    content: bytes
    rows_updated: int
    unmatched_updates: list[BidUpdate]


def optimized_filename(original_filename: str) -> str:
    """Append ``_Optimized`` before the ``.xlsx`` extension.

    ``CheapTicketsDealM_FIOAD_US.xlsx`` -> ``CheapTicketsDealM_FIOAD_US_Optimized.xlsx``.
    """
    stem = (
        original_filename[: -len(_XLSX_EXT)]
        if original_filename.lower().endswith(_XLSX_EXT)
        else original_filename
    )
    return f"{stem}{_EXPORT_SUFFIX}{_XLSX_EXT}"


def _text(value: object) -> str:
    """Trimmed string form of a cell value (``""`` for blanks)."""
    return "" if value is None else str(value).strip()


def _header_key(value: object) -> str:
    return _text(value).lower()


def _match_key(provider: str, origin: str, destination: str) -> tuple[str, str, str]:
    """Case-insensitive match key. Provider/route codes are compared upper-cased."""
    return (provider.strip().upper(), origin.strip().upper(), destination.strip().upper())


class BidWorkbookExporter:
    """Rewrite a bid workbook's Override CPC column, preserving everything else."""

    def apply_updates(self, source: bytes, updates: Sequence[BidUpdate]) -> WorkbookExportResult:
        """Return a copy of ``source`` with the given Override-CPC updates applied.

        Rows are matched by ``(provider, origin, destination)``. Updates that
        match no row are returned in :attr:`WorkbookExportResult.unmatched_updates`
        rather than raising — one bad route never fails the whole export.
        """
        # Not read-only and not data_only: keeps styles, merged cells, freeze
        # panes, named ranges, data validation, and formula strings intact.
        workbook = load_workbook(BytesIO(source))
        try:
            sheet = self._search_terms_sheet(workbook)
            header = [_header_key(cell.value) for cell in next(sheet.iter_rows(max_row=1))]
            col_provider = _column_index(header, "provider code")
            col_origin = _column_index(header, "origin")
            col_destination = _column_index(header, "destination")
            col_cpc = _column_index(header, "override cpc")
            if col_origin is None or col_destination is None or col_cpc is None:
                raise ParseError(
                    "bid workbook missing required columns: Origin, Destination, Override CPC"
                )

            mode_provider = self._mode_provider(workbook)
            pending = {_match_key(u.provider_code, u.origin, u.destination): u for u in updates}
            applied: set[tuple[str, str, str]] = set()

            for row in sheet.iter_rows(min_row=2):
                origin = _text(row[col_origin].value)
                destination = _text(row[col_destination].value)
                if not origin and not destination:
                    continue  # blank spacer row

                row_provider = _text(row[col_provider].value) if col_provider is not None else ""
                provider = row_provider or mode_provider
                key = _match_key(provider, origin, destination)
                update = pending.get(key)
                if update is None:
                    continue

                # The ONLY mutation performed anywhere in the workbook. Setting
                # ``.value`` leaves the cell's style/number-format untouched.
                row[col_cpc].value = round(float(update.new_cpc), 2)
                applied.add(key)

            content = self._dump(workbook)
            unmatched = [update for key, update in pending.items() if key not in applied]
            return WorkbookExportResult(
                content=content, rows_updated=len(applied), unmatched_updates=unmatched
            )
        finally:
            workbook.close()

    @staticmethod
    def _search_terms_sheet(workbook: Workbook) -> Worksheet:
        if SEARCH_TERMS_SHEET in workbook.sheetnames:
            return workbook[SEARCH_TERMS_SHEET]
        return workbook.worksheets[0]

    @staticmethod
    def _mode_provider(workbook: Workbook) -> str:
        """Provider Code from the Mode sheet (the fallback for blank route rows)."""
        if MODE_SHEET not in workbook.sheetnames:
            return ""
        for row in workbook[MODE_SHEET].iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            if _text(row[0]).lower() == "provider code" and len(row) > 1:
                return _text(row[1])
        return ""

    @staticmethod
    def _dump(workbook: Workbook) -> bytes:
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()


def _column_index(header: list[str], name: str) -> int | None:
    return header.index(name) if name in header else None
