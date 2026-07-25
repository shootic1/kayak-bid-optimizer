"""Parser for KAYAK bid workbooks (Search Terms + Mode sheets).

Read-only: this never modifies the workbook. Understands the real KAYAK bid file
structure documented in Phase 3A — provider is taken from the ``Mode`` sheet
(the Search Terms ``Provider Code`` column is only populated on the first row).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.importers.base import ParseError
from app.optimizer.matching import is_iata


@dataclass(frozen=True)
class ParsedBidRoute:
    row_index: int
    provider_code: str
    origin: str
    destination: str
    excluded: bool
    override_cpc: float | None
    origin_is_iata: bool
    destination_is_iata: bool


@dataclass(frozen=True)
class ParsedBidFile:
    provider_code: str
    mode: str
    routes: list[ParsedBidRoute]


def _to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def _cell(row: tuple[object, ...], index: int | None) -> object:
    if index is None or index >= len(row):
        return None
    return row[index]


class BidFileParser:
    """Parse a bid workbook into a :class:`ParsedBidFile`."""

    def parse(self, path: Path) -> ParsedBidFile:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # openpyxl raises assorted errors on bad files
            raise ParseError(f"could not read bid workbook: {exc}") from exc

        try:
            mode, provider_from_mode = self._read_mode_sheet(workbook)
            routes = self._read_search_terms(workbook, provider_from_mode)
            provider = provider_from_mode or (routes[0].provider_code if routes else "")
            return ParsedBidFile(provider_code=provider, mode=mode or "Full", routes=routes)
        finally:
            workbook.close()

    @staticmethod
    def _read_mode_sheet(workbook: object) -> tuple[str, str]:
        wb = workbook  # typed loosely to avoid openpyxl type noise
        if "Mode" not in wb.sheetnames:  # type: ignore[attr-defined]
            return "Full", ""
        mode = "Full"
        provider = ""
        for row in wb["Mode"].iter_rows(values_only=True):  # type: ignore[index]
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip().lower()
            value = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            if key == "mode" and value:
                mode = value
            elif key == "provider code" and value:
                provider = value
        return mode, provider

    def _read_search_terms(self, workbook: object, provider_default: str) -> list[ParsedBidRoute]:
        wb = workbook
        sheet = (
            wb["Search Terms"]  # type: ignore[index]
            if "Search Terms" in wb.sheetnames  # type: ignore[attr-defined]
            else wb.worksheets[0]  # type: ignore[attr-defined]
        )
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ParseError("bid workbook 'Search Terms' sheet is empty")

        norm = [str(h).strip().lower() if h is not None else "" for h in header]

        def col(name: str) -> int | None:
            return norm.index(name) if name in norm else None

        ci_provider = col("provider code")
        ci_origin = col("origin")
        ci_destination = col("destination")
        ci_excluded = col("excluded")
        ci_cpc = col("override cpc")
        if ci_origin is None or ci_destination is None or ci_cpc is None:
            raise ParseError(
                "bid workbook missing required columns: Origin, Destination, Override CPC"
            )

        parsed: list[ParsedBidRoute] = []
        for line_no, row in enumerate(rows, start=2):
            origin_raw = _cell(row, ci_origin)
            dest_raw = _cell(row, ci_destination)
            if (origin_raw is None or origin_raw == "") and (dest_raw is None or dest_raw == ""):
                continue  # blank row
            origin = str(origin_raw).strip() if origin_raw is not None else ""
            destination = str(dest_raw).strip() if dest_raw is not None else ""

            provider_raw = _cell(row, ci_provider)
            provider = str(provider_raw).strip() if provider_raw else provider_default

            excluded_raw = _cell(row, ci_excluded)
            excluded = str(excluded_raw).strip().lower() == "true" if excluded_raw else False

            parsed.append(
                ParsedBidRoute(
                    row_index=line_no,
                    provider_code=provider,
                    origin=origin,
                    destination=destination,
                    excluded=excluded,
                    override_cpc=_to_float(_cell(row, ci_cpc)),
                    origin_is_iata=is_iata(origin),
                    destination_is_iata=is_iata(destination),
                )
            )
        return parsed
