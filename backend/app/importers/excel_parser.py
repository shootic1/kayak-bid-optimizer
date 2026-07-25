"""Excel (.xlsx) parser built on openpyxl."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.domain.enums import FileType
from app.importers.base import BaseParser, ParsedTable, ParseError


class ExcelParser(BaseParser):
    """Parse the first worksheet of an .xlsx workbook."""

    file_type = FileType.XLSX

    def parse(self, path: Path) -> ParsedTable:
        try:
            workbook = load_workbook(filename=path, read_only=True, data_only=True)
        except Exception as exc:  # openpyxl raises assorted errors on bad files
            raise ParseError(f"could not read Excel file: {exc}") from exc

        try:
            sheet = workbook.worksheets[0]
            rows_iter = sheet.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                return ParsedTable(headers=[], rows=[])

            headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
            raw_rows = [list(row) for row in rows_iter]
            return ParsedTable(headers=headers, rows=self._build_rows(headers, raw_rows))
        finally:
            workbook.close()
